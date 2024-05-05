import openpyxl
from openpyxl.utils import get_column_letter

def sort_sinhala_words(excel_file, sheet_name, column_index):
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[sheet_name]

    # Get all the Sinhala words from the specified column
    sinhala_words = [cell.value for cell in sheet[get_column_letter(column_index)] if cell.value]

    # Sort the Sinhala words alphabetically
    sorted_sinhala_words = sorted(sinhala_words, key=lambda x: x.lower())

    # Write the sorted words back to the Excel file
    for i, word in enumerate(sorted_sinhala_words, start=1):
        sheet[get_column_letter(column_index) + str(i)] = word

    # Save the changes
    wb.save(excel_file)

    print("Sinhala words sorted and saved successfully.")

# Example usage
if __name__ == "__main__":
    excel_file = "words.xlsx"
    sheet_name = "Sheet1"
    column_index = 1

    sort_sinhala_words(excel_file, sheet_name, column_index)
